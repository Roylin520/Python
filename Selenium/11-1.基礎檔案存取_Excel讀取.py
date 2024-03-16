# pip install openpyxl 安裝讀取excel
import openpyxl
path = r'C:\Users\Roy\Documents\git\Python\Selenium\samples\excel.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)  # data_only=True 表示讀值，不讀公式

wk = wb.sheetnames
print(wk)  # 印出所有工作表名稱
print(wk[0])  # 印出索引0的工作表名稱
print(wk[1])  # 印出索引1的工作表名稱

# 獲取活動工作表
active_sheet = wb.active
print(active_sheet)

# 獲取第一個工作表
sheet1 = wb[wk[0]]  # 指定索引0

# 獲取第二個工作表
sheet2 = wb['部門']   # 指定工作表名稱

# 打印工作表相關信息，工作表名稱、資料最大筆數、最大欄位數
print(f"第一個工作表名稱：{sheet1.title}, 資料最大筆數：{sheet1.max_row}, 最大欄位數：{sheet1.max_column}")
print(f"第二個工作表名稱：{sheet2.title}, 資料最大筆數：{sheet2.max_row}, 最大欄位數：{sheet2.max_column}")


# %%
# 讀取儲存格內容
print(sheet1['A1'].value)
print(sheet1.cell(1, 1).value)
print(sheet2['B1'].value)
print(sheet2.cell(2, 2).value)